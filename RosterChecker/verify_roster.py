#!/usr/bin/env python3
"""
花名册核对脚本 (支持 --json 输出供 App 使用)
用法: python3 verify_roster.py <contact_list.pdf> <花名册.xlsx> [--json] [--output 结果.xlsx]
"""

import sys, os, re, json
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def parse_contact_list(pdf_path):
    skip = {'allstaff@290.com.hk','executivecommittee@290.com.hk','staff@fortune3369.com.hk',
            'staff@ffc.com.hk','assetmgmt@290.com.hk','szstaff@290.com.hk',
            'staff@fortune-finance.com.hk','finance@290.com.hk','capops@290.com.hk',
            'comsec@290.com.hk','legal@290.com.hk','compliance@290.com.hk',
            'edo@290.com.hk','ai@290.com.hk'}
    records = {}
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text: continue
            for line in text.split('\n'):
                m = re.search(r'([\w.+-]+@[\w.-]+\.\w+)', line)
                if not m: continue
                email = m.group(1).lower()
                if email in skip: continue
                prefix = line[:m.start()].strip()
                tokens = prefix.split()
                phones = []
                for t in tokens:
                    clean = t.replace(' ','').replace('-','').replace('+','').replace('(','').replace(')','')
                    if re.match(r'^\d{7,15}$', clean): phones.append(t)
                dl, mob = '', ''
                if len(phones) >= 2: dl, mob = phones[0], phones[-1]
                elif len(phones) == 1:
                    if '-' in phones[0]: dl = phones[0]
                    else: mob = phones[0]
                records[email] = {
                    'name': ' '.join(t for t in tokens if t not in phones and t != 'P'),
                    'direct_line': dl, 'mobile': mob
                }
    return records

def parse_roster(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['在职人员']
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    def fc(*keys):
        for k in keys:
            if k in header: return header.index(k)
        return None
    ni = fc('姓名'); ei = fc('工作邮箱'); pi = fc('直拨电话/移动电话')
    mi = fc('手机号码'); di = fc('部门'); si = fc('人员状态')
    if None in (ni,ei,pi,mi,di,si):
        raise ValueError("花名册缺少必要列")
    records = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[si] != '在职': continue
        email = (row[ei] or '').strip().lower()
        if not email: continue
        records[email] = {
            'name': (row[ni] or '').strip(),
            'direct_line': (row[pi] or '').strip(),
            'mobile': (row[mi] or '').strip(),
            'dept': (row[di] or '').strip()
        }
    return records

def norm_phone(p):
    p = p.replace(' ','').replace('-','').replace('+','').replace('(','').replace(')','')
    return re.sub(r'^(852|86)', '', p)

def phones_match(a, b):
    na, nb = norm_phone(a), norm_phone(b)
    if na == nb: return True
    if na and nb:
        if na.endswith(nb) or nb.endswith(na): return True
    return False

def compare(roster_data, cl_data):
    errors, only_roster, only_cl = [], [], []
    for email, ex in roster_data.items():
        if email not in cl_data:
            only_roster.append(ex); continue
        cl = cl_data[email]
        has_err = False
        err = {'name':ex['name'],'email':email,'dept':ex['dept'],
               'dl_excel':ex['direct_line'],'dl_cl':'','mob_excel':ex['mobile'],'mob_cl':cl['mobile']}
        if cl['direct_line']:
            err['dl_cl'] = cl['direct_line']
            if not phones_match(ex['direct_line'], cl['direct_line']): has_err = True
        else:
            err['dl_cl'] = cl['mobile']
            if cl['mobile'] and not phones_match(ex['direct_line'], cl['mobile']): has_err = True
        if cl['mobile'] and not phones_match(ex['mobile'], cl['mobile']): has_err = True
        if has_err: errors.append(err)
    for e in set(cl_data.keys()) - set(roster_data.keys()):
        only_cl.append(cl_data[e])
    return errors, only_roster, only_cl

def write_xlsx(errors, only_roster, only_cl, output_path):
    wb = openpyxl.Workbook()
    hf = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    hfw = Font(bold=True, size=11, color='FFFFFF')
    tb = Border(left=Side(style='thin'),right=Side(style='thin'),
                top=Side(style='thin'),bottom=Side(style='thin'))
    def sh(ws, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hfw; cell.fill = hf
            cell.alignment = Alignment(horizontal='center'); cell.border = tb
    def sd(ws, mr, mc):
        for r in range(2, mr+1):
            for c in range(1, mc+1):
                cell = ws.cell(row=r, column=c)
                cell.border = tb; cell.alignment = Alignment(horizontal='center')

    ws1 = wb.active; ws1.title = '信息不一致'
    h1 = ['姓名','邮箱','部门','直拨电话(花名册)','直拨电话(CL)','手机(花名册)','手机(CL)']
    sh(ws1, h1)
    for i, e in enumerate(errors, 2):
        for j, k in enumerate(['name','email','dept','dl_excel','dl_cl','mob_excel','mob_cl']):
            ws1.cell(row=i, column=j+1, value=e[k])
    sd(ws1, len(errors)+1, len(h1))

    ws2 = wb.create_sheet('花名册独有(CL无)')
    h2 = ['姓名','邮箱','部门','直拨电话','手机号码']
    sh(ws2, h2)
    for i, r in enumerate(only_roster, 2):
        for j, k in enumerate(['name','email','dept','direct_line','mobile']):
            ws2.cell(row=i, column=j+1, value=r.get(k,''))
    sd(ws2, len(only_roster)+1, len(h2))

    ws3 = wb.create_sheet('ContactList独有(花名册无)')
    h3 = ['姓名(英文)','邮箱','直拨电话(CL)','手机(CL)']
    sh(ws3, h3)
    for i, r in enumerate(only_cl, 2):
        for j, k in enumerate(['name','email','direct_line','mobile']):
            ws3.cell(row=i, column=j+1, value=r.get(k,''))
    sd(ws3, len(only_cl)+1, len(h3))

    for ws, cols in [(ws1,'ABCDEFG'),(ws2,'ABCDE'),(ws3,'ABCD')]:
        for c in cols: ws.column_dimensions[c].width = 22
    ws1.column_dimensions['B'].width = 36; ws3.column_dimensions['B'].width = 36

    wb.save(output_path)

# ── 主入口 ──
if __name__ == '__main__':
    json_mode = '--json' in sys.argv
    args = [a for a in sys.argv[1:] if not a.startswith('--') or a == '--json']
    args = [a for a in args if a != '--json']

    if len(args) < 2:
        print("用法: python3 verify_roster.py <contact_list.pdf> <花名册.xlsx> [--json] [--output 结果.xlsx]")
        sys.exit(1)

    pdf_path, xlsx_path = args[0], args[1]
    output_path = None
    for i, a in enumerate(sys.argv):
        if a == '--output' and i+1 < len(sys.argv):
            output_path = sys.argv[i+1]

    cl_data = parse_contact_list(pdf_path)
    roster_data = parse_roster(xlsx_path)
    errors, only_roster, only_cl = compare(roster_data, cl_data)

    if json_mode:
        result = {
            'summary': {
                'cl_count': len(cl_data),
                'roster_count': len(roster_data),
                'error_count': len(errors),
                'only_roster_count': len(only_roster),
                'only_cl_count': len(only_cl)
            },
            'errors': errors,
            'only_roster': [{'name':r['name'],'email':r.get('email',''),'dept':r['dept'],
                             'direct_line':r['direct_line'],'mobile':r['mobile']} for r in only_roster],
            'only_cl': [{'name':r['name'],'email':r.get('email',''),'direct_line':r['direct_line'],
                         'mobile':r['mobile']} for r in only_cl]
        }
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        out = output_path or '核对结果.xlsx'
        write_xlsx(errors, only_roster, only_cl, out)
        print(f"信息不一致: {len(errors)} 人")
        print(f"花名册独有: {len(only_roster)} 人")
        print(f"ContactList独有: {len(only_cl)} 人")
        print(f"结果已保存至: {out}")
